import csv
#note: need openpyxl module. Use sudo pip3 install openpyxl to install
import openpyxl
import datetime
from openpyxl import Workbook


xAccelRaw = []
xAccelCal = []
datas = []
time = []
timediff = []
velocity = []
velocity.insert(0, 0)

with open('/media/gcs-14/D606-E03A/DATALOG.CSV', 'r') as csv_file:
    csv_reader = csv.reader(csv_file)

    for line in csv_reader:
        datas.append(line)

for x in range(len(datas)):
    if datas[x][0] != '':
        time.append(datas[x][0])
        timediff.append(((int(datas[x][0])-(int(datas[0][0])-1))/1000))
        xAccelRaw.append(datas[x][1])
        cal = float(xAccelRaw[x])*(-1.085)-26.272
        xAccelCal.append(cal)
    else:
        break

xAccelCal.insert(0, 0)
timediff.insert(0, 0)

for x in range(1, len(xAccelCal)):
    velocity.append(float(velocity[x-1]) + (float(timediff[x])-float(timediff[x-1]))*(float(xAccelCal[x])+ float(xAccelCal[x-1]))/2)

for x in range(len(xAccelCal)):
    print(str(timediff[x]) + '\t' + str(velocity[x]))
csv_file.close()

book = openpyxl.load_workbook('my.xlsx')

SummarySheet = book.get_sheet_by_name('Summary')

date = str(datetime.date.today())

TrialSheet = book.create_sheet(date + "_0")

summaryVelocities = []

summaryVelocities.append(SummarySheet.cell(row=4, column=2).value)
summaryVelocities.append(SummarySheet.cell(row=5, column=2).value)
summaryVelocities.append(SummarySheet.cell(row=6, column=2).value)
summaryVelocities.append(SummarySheet.cell(row=7, column=2).value)
summaryVelocities.append(SummarySheet.cell(row=8, column=2).value)

maxTrialVelocity = max(velocity)

bump = True

for x in range(len(summaryVelocities)):
    if summaryVelocities[x]==0:
        summaryVelocities.insert(x, maxTrialVelocity)
        SummarySheet.cell(row=x+4, column = 2).value = maxTrialVelocity
        bump = False
        break

if bump==True:
    del summaryVelocities[0]
    summaryVelocities.append(maxTrialVelocity)
    SummarySheet.cell(row=4, column=2).value = summaryVelocities[0]
    SummarySheet.cell(row=5, column=2).value = summaryVelocities[1]
    SummarySheet.cell(row=6, column=2).value = summaryVelocities[2]
    SummarySheet.cell(row=7, column=2).value = summaryVelocities[3]
    SummarySheet.cell(row=8, column=2).value = summaryVelocities[4]

TrialSheet.cell(row=1, column=1).value = "Time (s)"
TrialSheet.cell(row=1, column=2).value = "XAccelRaw"
TrialSheet.cell(row=1, column=3).value = "xAccelCal (m/s^2)"
TrialSheet.cell(row=1, column=4).value = "Velocity (m/s)"
TrialSheet.cell(row=1, column=6).value = "Max Velocity (m/s):"
TrialSheet.cell(row=1, column=7).value = maxTrialVelocity

for row in zip(timediff, xAccelRaw, xAccelCal, velocity):
    TrialSheet.append(row)

book.save('my.xlsx')